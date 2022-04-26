#   ---------------------------------   Imports   -------------------------------------    #
import io
import pickle
import os
from io import BytesIO
import base64
import datetime as dt


from file_sending import  identify_lp_Test, Contacts_Test #Contacts, identify_lp, as
from google_auth_oauthlib.flow import Flow, InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from google.auth.transport.requests import Request
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import mimetypes
#   ---------------------------------   Functions   -------------------------------------    #


#   ---------------------------------   Create Service util' 1  --------------------------------------    #
def create_service(client_secret_file, api_name, api_version, *scopes):
    print(client_secret_file, api_name, api_version, scopes, sep='-')
    CLIENT_SECRET_FILE = client_secret_file
    API_SERVICE_NAME = api_name
    API_VERSION = api_version
    SCOPES = [scope for scope in scopes[0]]
    print(SCOPES)

    cred = None
    working_dir = os.getcwd()
    token_dir = 'token files'

    pickle_file = f'token_{API_SERVICE_NAME}_{API_VERSION}.pickle'

    ### Check if token dir exists first, if not, create the folder
    if not os.path.exists(os.path.join(working_dir, token_dir)):
        os.mkdir(os.path.join(working_dir, token_dir))

    if os.path.exists(os.path.join(working_dir, token_dir, pickle_file)):
        with open(os.path.join(working_dir, token_dir, pickle_file), 'rb') as token:
            cred = pickle.load(token)

    if not cred or not cred.valid:
        if cred and cred.expired and cred.refresh_token:
            cred.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            cred = flow.run_local_server()

        with open(os.path.join(working_dir, token_dir, pickle_file), 'wb') as token:
            pickle.dump(cred, token)

    try:
        service = build(API_SERVICE_NAME, API_VERSION, credentials=cred)
        print(API_SERVICE_NAME, 'service created successfully')
        return service
    except Exception as e:
        print(e)
        print(f'Failed to create service instance for {API_SERVICE_NAME}')
        os.remove(os.path.join(working_dir, token_dir, pickle_file))
        return None


#   ---------------------------------   Create Service util' 2  --------------------------------------    #
def construct_service(api_service, client_file):
    CLIENT_SERVICE_FILE = client_file
    try:
        if api_service == 'drive':
            API_Name = 'drive'
            API_Version = 'v3'
            Scopes = ['https://www.googleapis.com/auth/drive']
            return create_service(CLIENT_SERVICE_FILE, API_Name, API_Version, Scopes)
        elif api_service == 'gmail':
            API_Name = 'gmail'
            API_Version = 'v1'
            Scopes = ['https://mail.google.com/']
            return create_service(CLIENT_SERVICE_FILE, API_Name, API_Version, Scopes)
    except Exception as e:
        print(e)
        return None


#   ---------------------------------   Search Email by Parameters  --------------------------------------    #
def search_email(service, query_string, label_ids=[]):
    try:
        message_list_response = service.users().messages().list(
            userId='me',
            labelIds=label_ids,
            q=query_string
        ).execute()

        message_items = message_list_response.get('messages')
        nextPageToken = message_list_response.get('nextPageToken')

        while nextPageToken:
            message_list_response = service.users().messages().list(
                userId='me',
                labelIds=label_ids,
                q=query_string,
                pageToken=nextPageToken
            ).execute()
            message_items.extend(message_list_response.get('messages'))
            nextPageToken = message_list_response.get('nextPageToken')
        return message_items
    except Exception as e:
        return None


#   ---------------------------------   Retrieve Message Content  --------------------------------------    #
def get_message_detail(service, message_id, format='metadata', metadata_headers=[]):
    try:
        message_detail = service.users().messages().get(
            userId='me',
            id=message_id,
            format=format,
            metadataHeaders=metadata_headers
        ).execute()
        return message_detail
    except Exception as e:
        print(e)
        return None


#   ---------------------------------   Gmail API Credentials  --------------------------------------    #
def gmail_api_cerd():
    Client_file = 'client.json'
    API_Name = 'gmail'
    API_Version = 'v1'
    Scopes = ['https://mail.google.com/']

    return Client_file, API_Name, API_Version, Scopes


#   ---------------------------------   Initializing Veriables   -------------------------------------    #
def enitializig_veriabls(contacts):
    query_string = 'has:attachment'
    ecp_original_dir = 'Test Original'
    ecp_splited_dir = 'Test Splited'
    lp_emails = contacts
    current_date = dt.datetime.now().strftime("%d/%m/%Y")

    return query_string, ecp_original_dir, ecp_splited_dir, lp_emails, current_date


#   ---------------------------------   Download Emails   -------------------------------------    #
def download_email_attachments(email_messages, gmail_service):
    file_list = []
    # retrieving messages
    for email_message in email_messages:
        messageID = email_message['threadId']
        message_Subject = '(No Subject) ({0})'.format(messageID)
        message_Detail = get_message_detail(
            gmail_service,
            email_message['id'],
            format='full',
            metadata_headers=['parts'])
        message_Detail_Payload = message_Detail.get('payload')
        # retrieve headers from email
        for item in message_Detail_Payload['headers']:
            if item['name'] == 'Subject':
                if item['value']:
                    message_Subject = '{0} ({1})'.format(item['value'], messageID)
                else:
                    message_Subject = ' (No Subject) ({0})'.format(messageID)

        # extracting file from email
        if 'parts' in message_Detail_Payload:
            for msgPayload in message_Detail_Payload['parts']:
                mime_type = msgPayload['mimeType']
                file_name = msgPayload['filename']
                body = msgPayload['body']
                if 'attachmentId' in body:
                    attachment_id = body['attachmentId']
                    response = gmail_service.users().messages().attachments().get(
                        userId='me',
                        messageId=email_message['id'],
                        id=attachment_id
                    ).execute()
                    # encoding file data
                    file_data = base64.urlsafe_b64decode(
                        response.get('data').encode('UTF-8'))

                    # decoding file data for writing
                    fd = io.BytesIO(file_data)
                    with open(f"{ecp_original_dir}/{file_name}", 'wb') as file:  ## Open temporary file as bytes
                        file.write(fd.read())
                    # adding file name for extracted files list
                    file_list.append(f"{ecp_original_dir}/{file_name}")
    return file_list


#   ---------------------------------   Split files to separate worksheets   -------------------------------------    #
def file_split(file_name):
    split = {}
    wb = load_workbook(file_name)
    # splitting the workbook to different worksheets
    for ws in wb.sheetnames:
        new_wb = Workbook()
        new_wb.create_sheet(ws)
        # get max row & col
        mr = wb[ws].max_row
        mc = wb[ws].max_column
        # iterate over cells to copy data
        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                # reading cell value from source excel file
                c = wb[ws].cell(row=i, column=j)
                # writing the read value to destination excel file
                new_wb[ws].cell(row=i, column=j).value = c.value
        # removing default sheet
        del new_wb['Sheet']
        # saving the worksheet as a separate workbook
        new_wb.save(f"{ecp_splited_dir}/{ws} - {file_name[14:]}")
        split[f"{ws}"] = new_wb
    return split


#   ---------------------------------   Process the attached worksheet   -------------------------------------    #
def file_processing(file_names):
    files = {}
    for file_name in file_names:
        files[f"{file_name[14:]}"] = file_split(file_name)
    return files
    # for key in files.keys():
    #     if 'ECP' is in key:
    #
    #     if 'AMR' is in key:
    #


#   ---------------------------------   Send Relevant Worksheet to LP   -------------------------------------    #
def send_attachment(service, files_dict, date):
    email_message = f"Please see attached ECP daily report for {date}"
    for WB in files_dict:
        for WS in files_dict[WB]:
            print(files_dict[WB][WS])
            if WS in ['Elastum','eToroX','Hodleris','Paybis','Debex']:
                mimeMessage = MIMEMultipart()
                mimeMessage['to'] = identify_lp_Test(WS)
                mimeMessage['subject'] = f"ECP daily report for {date}"
                mimeMessage.attach(MIMEText(email_message,'plain'))
                # content_type, encoding = mimetypes.guess_type()
                # content_type= 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                # main_type, sub_type = content_type.split('/',1)
                file_name = f"{WS} - {WB}"
                # f = openpyxl.load_workbook(io.BytesIO(save_virtual_workbook(files_dict[attachment])))
                myfile = MIMEBase('application', 'vnd.ms-excel')
                myfile.set_payload(save_virtual_workbook(files_dict[WB][WS]))
                myfile.add_header('Content-Disposition', 'attachment', filename=file_name)
                encoders.encode_base64(myfile)
                mimeMessage.attach(myfile)
                raw_string = base64.urlsafe_b64encode(mimeMessage.as_bytes()).decode()
                service.users().messages().send(
                    userId='me',
                    body={'raw': raw_string}).execute()






if __name__ == '__main__':
    #   --------------------------------------------------------------------------------------    #
    Client_file, API_Name, API_Version, Scopes = gmail_api_cerd()
    #   --------------------------------------------------------------------------------------    #
    gmail_service = construct_service('gmail', Client_file)
    #   --------------------------------------------------------------------------------------    #
    query_string, ecp_original_dir, ecp_splited_dir, lp_emails, current_date = enitializig_veriabls(Contacts_Test)
    #   --------------------------------------------------------------------------------------    #
    email_messages = search_email(gmail_service, query_string, ['INBOX'])
    #   --------------------------------------------------------------------------------------    #
    file_names = download_email_attachments(email_messages, gmail_service)
    #   --------------------------------------------------------------------------------------    #
    files = file_processing(file_names)
    # files_dict = list(files.values())[0]
    #   --------------------------------------------------------------------------------------    #
    send_attachment(gmail_service, files, current_date)
